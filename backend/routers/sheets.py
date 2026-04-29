import json
import os
import uuid
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, Query, Header
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timezone

from database import get_db
from config import UPLOAD_DIR, COLUMN_LETTERS
from models.knowledge import Sheet
from routers.auth import get_user_from_token

router = APIRouter(prefix="/api/sheets", tags=["表格"])


# ========== Pydantic Schemas ==========

class SheetCreate(BaseModel):
    name: str
    mode: str = "manual"  # manual, learning, semi-auto, full-auto


class SheetUpdate(BaseModel):
    name: Optional[str] = None
    data_json: Optional[str] = None  # JSON字符串格式的表格数据
    mode: Optional[str] = None


class SheetResponse(BaseModel):
    id: int
    name: str
    data_json: str
    mode: str
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# ========== 工具函数 ==========

def get_token_from_header(authorization: str) -> str:
    """从Authorization header提取token"""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="缺少Bearer token")
    return authorization.replace("Bearer ", "")


# ========== 路由 ==========

@router.get("")
def list_sheets(
    authorization: str = Header(default=""),
    db: Session = Depends(get_db),
):
    """获取当前用户的所有表格列表"""
    token = get_token_from_header(authorization)
    user = get_user_from_token(token, db)
    sheets = db.query(Sheet).filter(Sheet.user_id == user.id).order_by(Sheet.updated_at.desc()).all()
    return {
        "success": True,
        "data": [SheetResponse.model_validate(s).model_dump() for s in sheets],
        "error": "",
    }


@router.post("")
def create_sheet(
    req: SheetCreate,
    authorization: str = Header(default=""),
    db: Session = Depends(get_db),
):
    """创建新表格"""
    token = get_token_from_header(authorization)
    user = get_user_from_token(token, db)

    # 新建表格，数据初始化为空数组
    sheet = Sheet(
        user_id=user.id,
        name=req.name,
        data_json="[]",
        mode=req.mode,
    )
    db.add(sheet)
    db.commit()
    db.refresh(sheet)

    return {
        "success": True,
        "data": SheetResponse.model_validate(sheet).model_dump(),
        "error": "",
    }


@router.get("/{sheet_id}")
def get_sheet(
    sheet_id: int,
    authorization: str = Header(default=""),
    db: Session = Depends(get_db),
):
    """获取单个表格的完整数据"""
    token = get_token_from_header(authorization)
    user = get_user_from_token(token, db)

    sheet = db.query(Sheet).filter(Sheet.id == sheet_id, Sheet.user_id == user.id).first()
    if not sheet:
        return {"success": False, "data": None, "error": "表格不存在或无权限访问"}

    return {
        "success": True,
        "data": SheetResponse.model_validate(sheet).model_dump(),
        "error": "",
    }


@router.put("/{sheet_id}")
def update_sheet(
    sheet_id: int,
    req: SheetUpdate,
    authorization: str = Header(default=""),
    db: Session = Depends(get_db),
):
    """更新表格数据"""
    token = get_token_from_header(authorization)
    user = get_user_from_token(token, db)

    sheet = db.query(Sheet).filter(Sheet.id == sheet_id, Sheet.user_id == user.id).first()
    if not sheet:
        return {"success": False, "data": None, "error": "表格不存在或无权限访问"}

    if req.name is not None:
        sheet.name = req.name
    if req.data_json is not None:
        sheet.data_json = req.data_json
    if req.mode is not None:
        sheet.mode = req.mode
    sheet.updated_at = datetime.now(timezone.utc)

    db.commit()
    db.refresh(sheet)

    return {
        "success": True,
        "data": SheetResponse.model_validate(sheet).model_dump(),
        "error": "",
    }


@router.post("/import")
async def import_file(
    file: UploadFile = File(...),
    authorization: str = Header(default=""),
    db: Session = Depends(get_db),
):
    """
    导入文件（xlsx/csv/txt），解析并返回18列结构化数据
    
    解析逻辑：
    - 读取上传文件
    - 提取A-E列（需求侧：序号、设备名称、招标/需求参数、单位、数量）
    - 返回完整的18列结构，F-R列初始为空
    """
    token = get_token_from_header(authorization)
    user = get_user_from_token(token, db)

    # 验证文件类型
    allowed_extensions = {".xlsx", ".xls", ".csv", ".txt"}
    file_ext = os.path.splitext(file.filename)[1].lower()
    if file_ext not in allowed_extensions:
        return {
            "success": False,
            "data": None,
            "error": f"不支持的文件格式: {file_ext}，支持: xlsx, csv, txt",
        }

    # 保存上传文件
    file_id = str(uuid.uuid4())
    safe_filename = f"{file_id}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, safe_filename)
    
    content = await file.read()
    with open(file_path, "wb") as f:
        f.write(content)

    try:
        # 解析文件 - 提取需求侧数据（A-E列）
        rows = _parse_import_file(file_path, file_ext)
    except Exception as e:
        return {"success": False, "data": None, "error": f"文件解析失败: {str(e)}"}

    # 返回rows数组（前端直接可用）
    return {
        "success": True,
        "data": rows,
        "error": "",
    }


@router.get("/{sheet_id}/export")
def export_sheet(
    sheet_id: int,
    authorization: str = Header(default=""),
    db: Session = Depends(get_db),
):
    """
    导出表格为格式化的xlsx文件
    
    格式要求：
    - 字体: SimSun 10pt
    - 自动列宽
    - 颜色分区: A-E列（需求区）浅蓝色, H-R列（产品区）浅绿色
    - F-G列留空作为分隔
    """
    token = get_token_from_header(authorization)
    user = get_user_from_token(token, db)

    sheet = db.query(Sheet).filter(Sheet.id == sheet_id, Sheet.user_id == user.id).first()
    if not sheet:
        return {"success": False, "data": None, "error": "表格不存在或无权限访问"}

    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        data = json.loads(sheet.data_json) if sheet.data_json else []
        
        wb = Workbook()
        ws = wb.active
        ws.title = sheet.name[:31]  # Excel sheet name max 31 chars

        # 定义样式
        header_font = Font(name="SimSun", size=10, bold=True)
        cell_font = Font(name="SimSun", size=10)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(name="SimSun", size=10, bold=True, color="FFFFFF")
        
        # 需求区（A-E）浅蓝色
        demand_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
        # 分隔区（F-G）浅灰色
        separator_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        # 产品区（H-R）浅绿色
        product_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # 列标题（A-R）
        columns_labels = ["序号", "设备名称", "招标/需求参数", "单位", "数量",
                          "", "",
                          "产品名称", "产品规格", "品牌", "产品型号", "厂家内部型号",
                          "数量", "单位", "产品单价", "总价", "备注", "不满足参数"]
        
        # 列宽设置
        col_widths = {
            "A": 8, "B": 20, "C": 40, "D": 8, "E": 10,
            "F": 4, "G": 4,
            "H": 20, "I": 35, "J": 12, "K": 18, "L": 18,
            "M": 10, "N": 8, "O": 12, "P": 14, "Q": 15, "R": 30,
        }

        # 设置列宽
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # 写表头
        col_letters = list("ABCDEFGHIJKLMNOPQR")
        for col_idx, (col_letter, label) in enumerate(zip(col_letters, columns_labels), start=1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        # 写数据行
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, col_letter in enumerate(col_letters):
                value = ""
                if isinstance(row_data, dict):
                    value = row_data.get(col_letter, "")
                elif isinstance(row_data, (list, tuple)) and col_idx < len(row_data):
                    value = row_data[col_idx]
                
                cell = ws.cell(row=row_idx, column=col_idx + 1, value=value)
                cell.font = cell_font
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = thin_border
                
                # 根据列区分颜色
                if col_letter in "ABCDE":
                    cell.fill = demand_fill
                elif col_letter in "FG":
                    cell.fill = separator_fill
                elif col_letter in "HIJKLMNOPQR":
                    cell.fill = product_fill

        # 保存到临时文件
        export_path = os.path.join(UPLOAD_DIR, f"export_{sheet_id}_{uuid.uuid4().hex[:8]}.xlsx")
        wb.save(export_path)

        return FileResponse(
            path=export_path,
            filename=f"{sheet.name}.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    except Exception as e:
        return {"success": False, "data": None, "error": f"导出失败: {str(e)}"}


# ========== 文件解析函数 ==========

def _parse_import_file(file_path: str, file_ext: str) -> list:
    """
    解析导入文件，提取需求侧数据（A-E列）
    返回18列格式的列表
    """
    rows = []

    if file_ext in (".xlsx", ".xls"):
        import pandas as pd
        df = pd.read_excel(file_path, dtype=str)
        rows = _df_to_18_columns(df)
    elif file_ext == ".csv":
        import pandas as pd
        # 尝试多种编码：UTF-8 BOM -> UTF-8 -> GBK -> GB18030
        encodings = ["utf-8-sig", "utf-8", "gbk", "gb18030", "latin-1"]
        df = None
        for enc in encodings:
            try:
                df = pd.read_csv(file_path, dtype=str, encoding=enc)
                break
            except (UnicodeDecodeError, UnicodeError):
                continue
        if df is None:
            df = pd.read_csv(file_path, dtype=str, encoding="utf-8", errors="replace")
        rows = _df_to_18_columns(df)
    elif file_ext == ".txt":
        with open(file_path, "r", encoding="utf-8", errors="replace") as f:
            lines = f.readlines()
        # 先检测分隔符：制表符优先，避免中文逗号误拆
        has_tab = any("\t" in line for line in lines[:10])
        for i, line in enumerate(lines):
            line = line.strip()
            if not line:
                continue
            if has_tab:
                parts = line.split("\t")
            else:
                # 智能CSV解析（处理引号内的逗号）
                parts = []
                current = ""
                in_quote = False
                for ch in line:
                    if ch == '"':
                        in_quote = not in_quote
                    elif ch == "," and not in_quote:
                        parts.append(current.strip())
                        current = ""
                    else:
                        current += ch
                parts.append(current.strip())
            
            row = {letter: "" for letter in COLUMN_LETTERS}
            # A=序号, B=设备名称, C=招标/需求参数, D=单位, E=数量
            row["A"] = str(i + 1)
            if len(parts) >= 1:
                row["B"] = parts[0].strip()
            if len(parts) >= 2:
                row["C"] = parts[1].strip()
            if len(parts) >= 3:
                row["D"] = parts[2].strip()
            if len(parts) >= 4:
                row["E"] = parts[3].strip()
            rows.append(row)

    return rows


def _df_to_18_columns(df) -> list:
    """将DataFrame转换为18列格式"""
    import pandas as pd
    rows = []
    # 获取列名列表
    columns = list(df.columns)
    
    for idx, row_data in df.iterrows():
        row = {letter: "" for letter in COLUMN_LETTERS}
        # A=序号
        row["A"] = str(idx + 1)
        
        # 尝试映射B-E列：从原始列中智能匹配
        values = [str(v) if pd.notna(v) else "" for v in row_data.values]
        
        if len(values) >= 1:
            row["B"] = values[0]  # 设备名称
        if len(values) >= 2:
            row["C"] = values[1]  # 招标/需求参数
        if len(values) >= 3:
            row["D"] = values[2]  # 单位
        if len(values) >= 4:
            row["E"] = values[3]  # 数量
        
        rows.append(row)
    
    return rows
